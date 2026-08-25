#!/usr/bin/env python3
"""
PPARgene v2.0 — 实验验证PPAR靶基因提取与DEG交叉分析 (Human-only)
==========================================================================
数据源: https://www.ppargene.org/download (CC-BY 4.0)
文献: Fang L et al., PPAR Research 2016:6042162, PMID: 27148361
      Qin J et al., bioRxiv 2025, DOI: 10.64898/2025.12.01.691485

⚠️ 物种过滤: PPARgene v2.0 含 7 物种 (human/mouse/rat/monkey/rabbit/
   bovine/pig)，本脚本仅提取 Species='human' 靶基因。
   非人类基因与 human DEG 交叉会产生假阳性——例如 CD36 在 PPARgene 中
   归属 mouse/rat，与 human DEG 交叉实为不同源基因。

输入:
  raw_data/PPARA.xlsx  — PPARα 实验验证靶基因 (含 Species 列)
  raw_data/PPARD.xlsx  — PPARβ/δ 实验验证靶基因
  raw_data/PPARG.xlsx  — PPARγ 实验验证靶基因
  DEG 文件 (deg_results_significant.tsv)

输出:
  results/PPARα_genes.txt         — PPARα unique gene symbols (human)
  results/PPARβ_δ_genes.txt       — PPARβ/δ unique gene symbols (human)
  results/PPARγ_genes.txt         — PPARγ unique gene symbols (human)
  results/PPARgene_union_genes.txt — all-subtype union (sorted, human)
  results/PPARgene_summary.json   — JSON summary
  results/PPARgene_full_analysis.tsv — per-gene annotation matrix
  results/PPARgene_DEG_overlap_18.tsv — DEG-overlapping genes (human)

修复记录:
  v1.1 (2026-06-20): 用有效数据行计数替代 ws.max_row-1，
                     避免 openpyxl trailing empty rows 导致行数虚高。
  v1.2 (2026-06-20): 添加物种过滤，仅保留 human 靶基因。
"""

import openpyxl
import json
import os
from pathlib import Path

# ── Paths ──────────────────────────────────────────────────────────
# All paths are resolved relative to this script's own location.
HERE = Path(__file__).resolve().parent
RAW_DIR = HERE / "raw_data"          # PPARα/β/δ .xlsx (PPARgene v2.0 download)
RES_DIR = HERE / "results"
RES_DIR.mkdir(parents=True, exist_ok=True)

# DEG file (Step1 standard pipeline output, copied into data/)
DEG_FILE = HERE / "data" / "deg_results_significant.tsv"

# ── Step 1: Load PPARgene data ─────────────────────────────────────
print("[1/4] Loading PPARgene verified target genes...")
SOURCE_FILES = [
    ("PPARA.xlsx", "PPARα"),
    ("PPARD.xlsx", "PPARβ/δ"),
    ("PPARG.xlsx", "PPARγ"),
]

all_symbols_upper = set()
gene_info = {}       # UPPER → {symbol, subtypes: set()}
subtype_sets = {}     # subtype → set(UPPER)
row_counts = {}

for fname, subtype in SOURCE_FILES:
    path = os.path.join(RAW_DIR, fname)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    subtype_sets[subtype] = set()

    # Count actual rows (skip trailing empty rows — openpyxl max_row
    # may include cells that were written then cleared, see e.g.
    # PPARD.xlsx max_row=150 but only 110 rows have data)
    data_rows = 0
    for r in range(2, ws.max_row + 1):
        g = ws.cell(r, 3).value    # Column C = Gene_symbol
        sp = ws.cell(r, 4).value   # Column D = Species
        if g and sp and str(g).strip() and str(sp).strip():
            if str(sp).strip().lower() != "human":
                continue  # v1.2: human-only filter
            g_orig = str(g).strip()
            g_upper = g_orig.upper()
            all_symbols_upper.add(g_upper)
            subtype_sets[subtype].add(g_upper)
            if g_upper not in gene_info:
                gene_info[g_upper] = {"symbol": g_orig, "subtypes": set()}
            gene_info[g_upper]["subtypes"].add(subtype)
            data_rows += 1
    row_counts[subtype] = data_rows

    # Save per-subtype gene list
    out = os.path.join(RES_DIR, f"{subtype.replace('/', '_')}_genes.txt")
    with open(out, "w") as f:
        for g in sorted(subtype_sets[subtype]):
            f.write(f"{gene_info[g]['symbol']}\n")
    print(f"  {subtype}: {len(subtype_sets[subtype])} genes ({row_counts[subtype]} rows)")

# Union
union_path = os.path.join(RES_DIR, "PPARgene_union_genes.txt")
with open(union_path, "w") as f:
    for g in sorted(all_symbols_upper):
        f.write(f"{gene_info[g]['symbol']}\n")
print(f"  Union: {len(all_symbols_upper)} unique genes")

# ── Step 2: Load DEG list ──────────────────────────────────────────
print(f"\n[2/4] Loading DEGs from {DEG_FILE}...")
deg_upper = set()
deg_map = {}  # UPPER → original symbol
with open(DEG_FILE) as f:
    header = f.readline().strip().split("\t")
    gene_col = header.index("Gene")
    for line in f:
        parts = line.strip().split("\t")
        if len(parts) > gene_col:
            g = parts[gene_col].strip('"').strip()
            gu = g.upper()
            deg_upper.add(gu)
            deg_map[gu] = g
print(f"  DEGs (FDR<0.05, |logFC|>1): {len(deg_upper)}")

# ── Step 3: Cross-reference ────────────────────────────────────────
print(f"\n[3/4] Cross-referencing...")
overlap = all_symbols_upper & deg_upper
print(f"  PPARgene ∩ DEG = {len(overlap)} genes")

# ── Step 4: Output ─────────────────────────────────────────────────
print(f"\n[4/4] Writing output files...")

# 4a. Full annotation matrix
matrix_path = os.path.join(RES_DIR, "PPARgene_full_analysis.tsv")
with open(matrix_path, "w") as f:
    f.write("Gene\tPPARα\tPPARβ/δ\tPPARγ\tIn_DEG\tIn_Overlap\n")
    for g in sorted(all_symbols_upper):
        info = gene_info[g]
        f.write(
            f"{info['symbol']}\t"
            f"{'Y' if g in subtype_sets['PPARα'] else ''}\t"
            f"{'Y' if g in subtype_sets['PPARβ/δ'] else ''}\t"
            f"{'Y' if g in subtype_sets['PPARγ'] else ''}\t"
            f"{'Y' if g in deg_upper else 'N'}\t"
            f"{'Y' if g in overlap else 'N'}\n"
        )
print(f"  {matrix_path}")

# 4b. DEG-overlap detail (human-only)
overlap_path = os.path.join(RES_DIR, "PPARgene_DEG_overlap_18.tsv")
with open(overlap_path, "w") as f:
    f.write("Gene\tPPARα\tPPARβ/δ\tPPARγ\n")
    for g in sorted(overlap):
        info = gene_info[g]
        f.write(
            f"{deg_map[g]}\t"
            f"{'Y' if g in subtype_sets['PPARα'] else ''}\t"
            f"{'Y' if g in subtype_sets['PPARβ/δ'] else ''}\t"
            f"{'Y' if g in subtype_sets['PPARγ'] else ''}\n"
        )
print(f"  {overlap_path}")

# 4c. JSON summary
summary = {
    "database": "PPARgene v2.0",
    "url": "https://www.ppargene.org/",
    "license": "CC-BY 4.0",
    "references": [
        "Fang L et al., PPAR Research 2016:6042162, PMID: 27148361",
        "Qin J et al., bioRxiv 2025, DOI: 10.64898/2025.12.01.691485"
    ],
    "subtypes": {
        st: {
            "unique_genes": len(subtype_sets[st]),
            "total_rows": row_counts[st]
        }
        for st in subtype_sets
    },
    "union_genes": len(all_symbols_upper),
    "deg_total": len(deg_upper),
    "ppargene_deg_overlap": len(overlap),
    "overlap_genes": [deg_map[g] for g in sorted(overlap)]
}
summary_path = os.path.join(RES_DIR, "PPARgene_summary.json")
with open(summary_path, "w") as f:
    json.dump(summary, f, indent=2, ensure_ascii=False)
print(f"  {summary_path}")

print("\n✅ Done. All outputs in", RES_DIR)
